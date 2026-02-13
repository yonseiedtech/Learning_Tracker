from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify
from app.decorators import auth_required, get_current_user
from app import firestore_dao as dao
from app.forms import SubjectForm
from datetime import datetime
import io
from openpyxl import Workbook, load_workbook

bp = Blueprint('subjects', __name__, url_prefix='/subjects')

MAX_FILE_SIZE = 100 * 1024 * 1024

ROLE_DISPLAY = {
    'instructor': '강사',
    'assistant': '조교',
    'student': '학생',
    'ta': 'TA',
    'auditor': '청강생'
}


def get_role_display(role):
    return ROLE_DISPLAY.get(role, role)


def has_subject_access(subject, user, roles=None):
    if roles is None:
        roles = ['instructor', 'assistant']
    if subject['instructor_id'] == user['uid']:
        return True
    member = dao.get_subject_member(subject['id'], user['uid'])
    if member and member['role'] in roles:
        return True
    return False


@bp.route('/')
@auth_required
def list_subjects():
    user = get_current_user()
    if user.get('role') == 'instructor':
        enrolled_subjects = dao.get_subjects_by_instructor(user['uid'])
        # Filter out deleted
        enrolled_subjects = [s for s in enrolled_subjects if not s.get('deleted_at')]
        all_subjects = []
    else:
        user_enrollments = dao.get_subject_enrollments_by_user(user['uid'])
        enrolled_ids = [e['subject_id'] for e in user_enrollments]
        enrolled_subjects = []
        for sid in enrolled_ids:
            s = dao.get_subject(sid)
            if s and not s.get('deleted_at') and s.get('is_visible', True):
                enrolled_subjects.append(s)
        # Get all visible subjects not enrolled in
        from app.firestore_dao import get_db
        all_subjects_raw = [doc.to_dict() | {'id': doc.id} for doc in get_db().collection('subjects').stream()]
        all_subjects = [
            s for s in all_subjects_raw
            if s['id'] not in enrolled_ids
            and not s.get('deleted_at')
            and s.get('is_visible', True)
        ]

    # Enrich subjects with course_count
    for s in enrolled_subjects:
        s['course_count'] = len(dao.get_courses_by_subject(s['id']))
    for s in all_subjects:
        s['course_count'] = len(dao.get_courses_by_subject(s['id']))

    # Enrich subjects with instructor data for templates
    for s in enrolled_subjects:
        dao.enrich_subject(s)
    for s in all_subjects:
        dao.enrich_subject(s)

    return render_template('subjects/list.html',
                          enrolled_subjects=enrolled_subjects,
                          all_subjects=all_subjects,
                          is_instructor=user.get('role') == 'instructor')


@bp.route('/create', methods=['GET', 'POST'])
@auth_required
def create():
    user = get_current_user()
    if user.get('role') != 'instructor':
        flash('강사만 과목을 생성할 수 있습니다.', 'danger')
        return redirect(url_for('subjects.list_subjects'))

    form = SubjectForm()
    if form.validate_on_submit():
        invite_code = dao.generate_invite_code('subjects')
        subject_id = dao.create_subject({
            'title': form.title.data,
            'description': form.description.data,
            'instructor_id': user['uid'],
            'invite_code': invite_code,
            'is_visible': True,
            'created_at': datetime.utcnow().isoformat()
        })

        dao.create_subject_member({
            'subject_id': subject_id,
            'user_id': user['uid'],
            'role': 'instructor'
        })
        flash('과목이 생성되었습니다.', 'success')
        return redirect(url_for('subjects.view', subject_id=subject_id))
    return render_template('subjects/create.html', form=form)


@bp.route('/<subject_id>')
@auth_required
def view(subject_id):
    subject = dao.get_subject(subject_id)
    if not subject:
        flash('해당 과목을 찾을 수 없습니다.', 'danger')
        return redirect(url_for('subjects.list_subjects'))

    user = get_current_user()

    if subject.get('deleted_at'):
        if not (user.get('role') == 'instructor' and subject['instructor_id'] == user['uid']):
            flash('해당 과목을 찾을 수 없습니다.', 'danger')
            return redirect(url_for('subjects.list_subjects'))

    if not subject.get('is_visible', True):
        if not (user.get('role') == 'instructor' and subject['instructor_id'] == user['uid']):
            flash('해당 과목은 비공개 상태입니다.', 'warning')
            return redirect(url_for('subjects.list_subjects'))

    courses = dao.get_courses_by_subject(subject_id)
    # Filter out deleted courses and sort
    courses = [c for c in courses if not c.get('deleted_at')]
    courses.sort(key=lambda c: (c.get('order_number') or 9999, c.get('week_number') or 9999, c.get('created_at', '')))

    is_enrolled = False
    if user.get('role') != 'instructor':
        subject_enrollment = dao.get_subject_enrollment(subject_id, user['uid'])
        is_enrolled = subject_enrollment is not None

    approved_enrollments = dao.get_subject_enrollments_by_subject(subject_id, status='approved')
    enrolled_count = len(approved_enrollments)
    pending_enrollments = dao.get_subject_enrollments_by_subject(subject_id, status='pending')
    pending_count = len(pending_enrollments)

    # Enrich subject with member counts
    members = dao.get_subject_members(subject_id)
    subject['instructor_count'] = len([m for m in members if m.get('role') == 'instructor'])
    subject['assistant_count'] = len([m for m in members if m.get('role') == 'assistant'])

    # Enrich courses with checkpoint_count and active_session
    for course in courses:
        checkpoints = dao.get_checkpoints_by_course(course['id'])
        course['checkpoint_count'] = len([cp for cp in checkpoints if not cp.get('deleted_at')])
        course['active_session'] = dao.get_active_session_for_course(course['id'])

    # Enrich subject with instructor and courses with instructor/subject data
    dao.enrich_subject(subject)
    dao.enrich_courses(courses)

    return render_template('subjects/view.html', subject=subject, courses=courses,
                          is_enrolled=is_enrolled, enrolled_count=enrolled_count,
                          pending_count=pending_count)


@bp.route('/<subject_id>/edit', methods=['GET', 'POST'])
@auth_required
def edit(subject_id):
    subject = dao.get_subject(subject_id)
    if not subject:
        flash('해당 과목을 찾을 수 없습니다.', 'danger')
        return redirect(url_for('subjects.list_subjects'))
    user = get_current_user()
    if not has_subject_access(subject, user):
        flash('접근 권한이 없습니다.', 'danger')
        return redirect(url_for('subjects.list_subjects'))

    form = SubjectForm(data=subject)
    if form.validate_on_submit():
        dao.update_subject(subject_id, {
            'title': form.title.data,
            'description': form.description.data
        })
        flash('과목이 수정되었습니다.', 'success')
        return redirect(url_for('subjects.view', subject_id=subject_id))
    return render_template('subjects/edit.html', form=form, subject=subject)


@bp.route('/<subject_id>/add-course', methods=['GET', 'POST'])
@auth_required
def add_course(subject_id):
    subject = dao.get_subject(subject_id)
    if not subject:
        flash('해당 과목을 찾을 수 없습니다.', 'danger')
        return redirect(url_for('subjects.list_subjects'))
    user = get_current_user()
    if not has_subject_access(subject, user):
        flash('접근 권한이 없습니다.', 'danger')
        return redirect(url_for('subjects.view', subject_id=subject_id))

    from app.forms import CourseForm
    from app.services.storage import upload_video, upload_material

    form = CourseForm()

    if form.validate_on_submit():
        existing_courses = dao.get_courses_by_subject(subject_id)
        order_num = form.order_number.data if form.order_number.data else len(existing_courses) + 1

        course_data = {
            'title': form.title.data,
            'description': form.description.data,
            'instructor_id': user['uid'],
            'subject_id': subject_id,
            'order_number': order_num,
            'session_type': form.session_type.data,
            'visibility': form.visibility.data,
            'video_url': form.video_url.data if form.session_type.data == 'video_external' else None,
            'assignment_description': form.assignment_description.data if form.session_type.data == 'assignment' else None,
            'quiz_time_limit': form.quiz_time_limit.data if form.session_type.data == 'quiz' else None,
            'quiz_pass_score': form.quiz_pass_score.data if form.session_type.data == 'quiz' else None,
            'invite_code': dao.generate_invite_code('courses'),
            'preparation_status': 'pending',
            'created_at': datetime.utcnow().isoformat()
        }

        if form.start_date.data:
            course_data['start_date'] = form.start_date.data
        if form.end_date.data:
            course_data['end_date'] = form.end_date.data
        if form.attendance_start.data:
            course_data['attendance_start'] = form.attendance_start.data
        if form.attendance_end.data:
            course_data['attendance_end'] = form.attendance_end.data
        if form.late_allowed.data and form.late_end.data:
            course_data['late_allowed'] = True
            course_data['late_end'] = form.late_end.data
        if form.assignment_due_date.data:
            course_data['assignment_due_date'] = form.assignment_due_date.data

        prereq_id = request.form.get('prerequisite_course_id')
        if prereq_id and form.visibility.data == 'prerequisite':
            course_data['prerequisite_course_id'] = prereq_id

        if form.session_type.data == 'video' and 'video_file' in request.files:
            video_file = request.files['video_file']
            if video_file and video_file.filename:
                file_content = video_file.read()
                if len(file_content) > MAX_FILE_SIZE:
                    flash('파일 크기가 100MB를 초과합니다.', 'danger')
                    return render_template('subjects/add_course.html', form=form, subject=subject)
                video_url = upload_video(file_content, video_file.filename)
                course_data['video_file_name'] = video_file.filename
                course_data['video_file_url'] = video_url
                course_data['preparation_status'] = 'ready'
        elif form.session_type.data == 'video_external' and form.video_url.data:
            course_data['preparation_status'] = 'ready'

        if form.session_type.data == 'material' and 'material_file' in request.files:
            material_file = request.files['material_file']
            if material_file and material_file.filename:
                file_content = material_file.read()
                if len(file_content) > MAX_FILE_SIZE:
                    flash('파일 크기가 100MB를 초과합니다.', 'danger')
                    return render_template('subjects/add_course.html', form=form, subject=subject)
                material_url = upload_material(file_content, material_file.filename)
                course_data['material_file_name'] = material_file.filename
                ext = material_file.filename.rsplit('.', 1)[-1].lower() if '.' in material_file.filename else ''
                course_data['material_file_type'] = ext
                course_data['material_file_url'] = material_url
                course_data['preparation_status'] = 'ready'

        if form.session_type.data == 'live_session':
            course_data['preparation_status'] = 'ready'

        dao.create_course(course_data)

        session_type_names = {
            'live_session': '라이브 세션',
            'video': '동영상 시청',
            'material': '학습 자료',
            'assignment': '과제 제출',
            'quiz': '퀴즈'
        }
        flash(f'{session_type_names.get(form.session_type.data, "세션")}이(가) 추가되었습니다.', 'success')
        return redirect(url_for('subjects.view', subject_id=subject_id))

    return render_template('subjects/add_course.html', form=form, subject=subject)


@bp.route('/<subject_id>/regenerate-code', methods=['POST'])
@auth_required
def regenerate_code(subject_id):
    subject = dao.get_subject(subject_id)
    if not subject:
        return jsonify({'success': False, 'message': '과목을 찾을 수 없습니다.'}), 404
    user = get_current_user()
    if not has_subject_access(subject, user):
        return jsonify({'success': False, 'message': '권한이 없습니다.'}), 403

    new_code = dao.generate_invite_code('subjects')
    dao.update_subject(subject_id, {'invite_code': new_code})
    return jsonify({'success': True, 'invite_code': new_code})


@bp.route('/<subject_id>/enroll', methods=['POST'])
@auth_required
def enroll_subject(subject_id):
    user = get_current_user()
    if user.get('role') == 'instructor':
        flash('강사는 과목 등록을 할 수 없습니다.', 'warning')
        return redirect(url_for('subjects.view', subject_id=subject_id))

    subject = dao.get_subject(subject_id)
    if not subject:
        flash('해당 과목을 찾을 수 없습니다.', 'danger')
        return redirect(url_for('subjects.list_subjects'))

    if subject.get('deleted_at') or not subject.get('is_visible', True):
        flash('해당 과목에 등록할 수 없습니다.', 'danger')
        return redirect(url_for('subjects.list_subjects'))

    existing_enrollment = dao.get_subject_enrollment(subject_id, user['uid'])

    if existing_enrollment:
        flash('이미 해당 과목에 등록되어 있습니다.', 'info')
        return redirect(url_for('subjects.view', subject_id=subject_id))

    dao.create_subject_enrollment({
        'subject_id': subject_id,
        'user_id': user['uid'],
        'status': 'approved',
        'created_at': datetime.utcnow().isoformat()
    })

    courses = dao.get_courses_by_subject(subject_id)
    courses = [c for c in courses if not c.get('deleted_at') and c.get('visibility') != 'private']
    enrolled_count = 0
    for course in courses:
        if not dao.is_enrolled(user['uid'], course['id']):
            dao.create_enrollment({
                'user_id': user['uid'],
                'course_id': course['id'],
                'created_at': datetime.utcnow().isoformat()
            })
            enrolled_count += 1

    flash(f'{subject["title"]} 과목에 등록되었습니다. ({enrolled_count}개 세션 자동 등록)', 'success')
    return redirect(url_for('subjects.view', subject_id=subject_id))


@bp.route('/<subject_id>/unenroll', methods=['POST'])
@auth_required
def unenroll_subject(subject_id):
    user = get_current_user()
    if user.get('role') == 'instructor':
        flash('강사는 과목 등록 취소를 할 수 없습니다.', 'warning')
        return redirect(url_for('subjects.view', subject_id=subject_id))

    subject = dao.get_subject(subject_id)
    if not subject:
        flash('해당 과목을 찾을 수 없습니다.', 'danger')
        return redirect(url_for('subjects.list_subjects'))

    subject_enrollment = dao.get_subject_enrollment(subject_id, user['uid'])

    if not subject_enrollment:
        flash('해당 과목에 등록되어 있지 않습니다.', 'warning')
        return redirect(url_for('subjects.view', subject_id=subject_id))

    dao.delete_subject_enrollment(subject_id, user['uid'])

    courses = dao.get_courses_by_subject(subject_id)
    for course in courses:
        enrollment = dao.get_enrollment(course['id'], user['uid'])
        if enrollment:
            # Delete the course enrollment - use dao method
            from app.firestore_dao import get_db
            get_db().collection('enrollments').document(enrollment['id']).delete()

    flash(f'{subject["title"]} 과목 등록이 취소되었습니다.', 'success')
    return redirect(url_for('subjects.list_subjects'))


@bp.route('/enroll-by-code', methods=['POST'])
@auth_required
def enroll_by_code():
    user = get_current_user()
    if user.get('role') == 'instructor':
        flash('강사는 과목 등록을 할 수 없습니다.', 'warning')
        return redirect(url_for('main.dashboard'))

    invite_code = request.form.get('invite_code', '').strip().upper()
    if not invite_code:
        flash('초대 코드를 입력하세요.', 'danger')
        return redirect(url_for('subjects.list_subjects'))

    subject = dao.get_subject_by_invite_code(invite_code)
    if subject and not subject.get('deleted_at') and subject.get('is_visible', True):
        existing_enrollment = dao.get_subject_enrollment(subject['id'], user['uid'])

        if existing_enrollment:
            flash('이미 해당 과목에 등록되어 있습니다.', 'info')
            return redirect(url_for('subjects.view', subject_id=subject['id']))

        dao.create_subject_enrollment({
            'subject_id': subject['id'],
            'user_id': user['uid'],
            'status': 'approved',
            'created_at': datetime.utcnow().isoformat()
        })

        courses = dao.get_courses_by_subject(subject['id'])
        courses = [c for c in courses if not c.get('deleted_at') and c.get('visibility') != 'private']
        enrolled_count = 0
        for course in courses:
            if not dao.is_enrolled(user['uid'], course['id']):
                dao.create_enrollment({
                    'user_id': user['uid'],
                    'course_id': course['id'],
                    'created_at': datetime.utcnow().isoformat()
                })
                enrolled_count += 1

        flash(f'{subject["title"]} 과목에 등록되었습니다. ({enrolled_count}개 세션 자동 등록)', 'success')
        return redirect(url_for('subjects.view', subject_id=subject['id']))

    flash('유효하지 않은 초대 코드입니다.', 'danger')
    return redirect(url_for('subjects.list_subjects'))


@bp.route('/<subject_id>/delete', methods=['POST'])
@auth_required
def delete_subject(subject_id):
    subject = dao.get_subject(subject_id)
    if not subject:
        flash('해당 과목을 찾을 수 없습니다.', 'danger')
        return redirect(url_for('subjects.list_subjects'))
    user = get_current_user()
    if not has_subject_access(subject, user):
        flash('접근 권한이 없습니다.', 'danger')
        return redirect(url_for('subjects.list_subjects'))

    dao.update_subject(subject_id, {'deleted_at': datetime.utcnow().isoformat()})
    flash(f'{subject["title"]} 과목이 삭제되었습니다.', 'success')
    return redirect(url_for('subjects.list_subjects'))


@bp.route('/<subject_id>/toggle-visibility', methods=['POST'])
@auth_required
def toggle_subject_visibility(subject_id):
    subject = dao.get_subject(subject_id)
    if not subject:
        flash('해당 과목을 찾을 수 없습니다.', 'danger')
        return redirect(url_for('subjects.list_subjects'))
    user = get_current_user()
    if not has_subject_access(subject, user):
        flash('접근 권한이 없습니다.', 'danger')
        return redirect(url_for('subjects.view', subject_id=subject_id))

    new_visibility = not subject.get('is_visible', True)
    dao.update_subject(subject_id, {'is_visible': new_visibility})
    status = '공개' if new_visibility else '비공개'
    flash(f'{subject["title"]} 과목이 {status}로 변경되었습니다.', 'success')
    return redirect(url_for('subjects.view', subject_id=subject_id))


@bp.route('/course/<course_id>/delete', methods=['POST'])
@auth_required
def delete_course(course_id):
    course = dao.get_course(course_id)
    if not course:
        flash('세션을 찾을 수 없습니다.', 'danger')
        return redirect(url_for('subjects.list_subjects'))
    user = get_current_user()
    if course['instructor_id'] != user['uid']:
        flash('접근 권한이 없습니다.', 'danger')
        return redirect(url_for('subjects.list_subjects'))

    subject_id = course.get('subject_id')
    dao.update_course(course_id, {'deleted_at': datetime.utcnow().isoformat()})
    flash(f'{course["title"]} 세션이 삭제되었습니다.', 'success')

    if subject_id:
        return redirect(url_for('subjects.view', subject_id=subject_id))
    return redirect(url_for('subjects.list_subjects'))


@bp.route('/course/<course_id>/toggle-visibility', methods=['POST'])
@auth_required
def toggle_course_visibility(course_id):
    course = dao.get_course(course_id)
    if not course:
        flash('세션을 찾을 수 없습니다.', 'danger')
        return redirect(url_for('subjects.list_subjects'))
    user = get_current_user()
    if course['instructor_id'] != user['uid']:
        flash('접근 권한이 없습니다.', 'danger')
        return redirect(url_for('subjects.list_subjects'))

    if course.get('visibility') == 'private':
        new_visibility = 'public'
        status = '공개'
    else:
        new_visibility = 'private'
        status = '비공개'
    dao.update_course(course_id, {'visibility': new_visibility})
    flash(f'{course["title"]} 세션이 {status}로 변경되었습니다.', 'success')

    if course.get('subject_id'):
        return redirect(url_for('subjects.view', subject_id=course['subject_id']))
    return redirect(url_for('courses.view', course_id=course_id))


@bp.route('/course/<course_id>/get', methods=['GET'])
@auth_required
def get_course(course_id):
    course = dao.get_course(course_id)
    if not course:
        return jsonify({'success': False, 'message': '세션을 찾을 수 없습니다.'}), 404
    user = get_current_user()
    if course['instructor_id'] != user['uid']:
        return jsonify({'success': False, 'message': '강사만 세션을 수정할 수 있습니다.'}), 403

    return jsonify({
        'success': True,
        'course': {
            'id': course['id'],
            'title': course.get('title', ''),
            'description': course.get('description', ''),
            'session_type': course.get('session_type', ''),
            'order_number': course.get('order_number') or course.get('week_number') or 1,
            'visibility': course.get('visibility', 'public'),
            'video_url': course.get('video_url', ''),
            'video_file_name': course.get('video_file_name', ''),
            'material_file_name': course.get('material_file_name', ''),
            'assignment_description': course.get('assignment_description', ''),
            'assignment_due_date': course.get('assignment_due_date', ''),
            'quiz_time_limit': course.get('quiz_time_limit', 30),
            'quiz_pass_score': course.get('quiz_pass_score', 60),
            'start_date': course.get('start_date', ''),
            'end_date': course.get('end_date', ''),
            'attendance_start': course.get('attendance_start', ''),
            'attendance_end': course.get('attendance_end', ''),
            'late_allowed': course.get('late_allowed', False),
            'late_end': course.get('late_end', ''),
        }
    })


@bp.route('/course/<course_id>/update', methods=['POST'])
@auth_required
def update_course(course_id):
    from app.services.storage import upload_video, upload_material

    course = dao.get_course(course_id)
    if not course:
        return jsonify({'success': False, 'message': '세션을 찾을 수 없습니다.'}), 404
    user = get_current_user()
    if course['instructor_id'] != user['uid']:
        return jsonify({'success': False, 'message': '강사만 세션을 수정할 수 있습니다.'}), 403

    try:
        data = request.form
        update_data = {}

        update_data['title'] = data.get('title', course.get('title', ''))
        update_data['description'] = data.get('description', course.get('description', ''))
        if data.get('order_number'):
            update_data['order_number'] = int(data.get('order_number'))
        update_data['visibility'] = data.get('visibility', course.get('visibility', 'public'))

        if course.get('session_type') == 'video_external':
            new_url = data.get('video_url', '')
            if new_url:
                update_data['video_url'] = new_url
                update_data['preparation_status'] = 'ready'

        if course.get('session_type') == 'video' and 'video_file' in request.files:
            video_file = request.files['video_file']
            if video_file and video_file.filename:
                file_content = video_file.read()
                if len(file_content) > MAX_FILE_SIZE:
                    return jsonify({'success': False, 'message': '파일 크기가 100MB를 초과합니다.'}), 400
                video_url = upload_video(file_content, video_file.filename)
                update_data['video_file_name'] = video_file.filename
                update_data['video_file_url'] = video_url
                update_data['preparation_status'] = 'ready'

        if course.get('session_type') == 'material' and 'material_file' in request.files:
            material_file = request.files['material_file']
            if material_file and material_file.filename:
                file_content = material_file.read()
                if len(file_content) > MAX_FILE_SIZE:
                    return jsonify({'success': False, 'message': '파일 크기가 100MB를 초과합니다.'}), 400
                material_url = upload_material(file_content, material_file.filename)
                update_data['material_file_name'] = material_file.filename
                ext = material_file.filename.rsplit('.', 1)[-1].lower() if '.' in material_file.filename else ''
                update_data['material_file_type'] = ext
                update_data['material_file_url'] = material_url
                update_data['preparation_status'] = 'ready'

        if course.get('session_type') == 'assignment':
            update_data['assignment_description'] = data.get('assignment_description', course.get('assignment_description', ''))
            if data.get('assignment_due_date'):
                update_data['assignment_due_date'] = data.get('assignment_due_date')

        if course.get('session_type') == 'quiz':
            update_data['quiz_time_limit'] = int(data.get('quiz_time_limit', 30))
            update_data['quiz_pass_score'] = int(data.get('quiz_pass_score', 60))

        if data.get('start_date'):
            update_data['start_date'] = data.get('start_date')
        if data.get('end_date'):
            update_data['end_date'] = data.get('end_date')
        if data.get('attendance_start'):
            update_data['attendance_start'] = data.get('attendance_start')
        if data.get('attendance_end'):
            update_data['attendance_end'] = data.get('attendance_end')
        if data.get('late_allowed') == 'on' and data.get('late_end'):
            update_data['late_allowed'] = True
            update_data['late_end'] = data.get('late_end')
        else:
            update_data['late_allowed'] = False
            update_data['late_end'] = None

        dao.update_course(course_id, update_data)

        return jsonify({'success': True, 'message': '세션이 성공적으로 수정되었습니다.'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/<subject_id>/members')
@auth_required
def members(subject_id):
    subject = dao.get_subject(subject_id)
    if not subject:
        flash('해당 과목을 찾을 수 없습니다.', 'danger')
        return redirect(url_for('subjects.list_subjects'))
    user = get_current_user()
    if not has_subject_access(subject, user):
        flash('접근 권한이 없습니다.', 'danger')
        return redirect(url_for('subjects.view', subject_id=subject_id))

    members_list = dao.get_subject_members(subject_id)

    member_data = {
        'instructors': [],
        'assistants': [],
        'students': []
    }

    for member in members_list:
        member_user = dao.get_user(member['user_id'])
        data = {'member': member, 'user': member_user}
        if member['role'] == 'instructor':
            member_data['instructors'].append(data)
        elif member['role'] == 'assistant':
            member_data['assistants'].append(data)
        else:
            member_data['students'].append(data)

    enrollments = dao.get_subject_enrollments_by_subject(subject_id)

    enrollment_data = {
        'approved': [],
        'pending': [],
        'rejected': []
    }

    for enrollment in enrollments:
        enrollment_user = dao.get_user(enrollment['user_id'])
        data = {'enrollment': enrollment, 'user': enrollment_user}
        status = enrollment.get('status', 'approved')
        if status in enrollment_data:
            enrollment_data[status].append(data)
        else:
            enrollment_data['approved'].append(data)

    return render_template('subjects/members.html',
                          subject=subject,
                          member_data=member_data,
                          enrollment_data=enrollment_data)


@bp.route('/<subject_id>/members/add', methods=['POST'])
@auth_required
def add_member(subject_id):
    subject = dao.get_subject(subject_id)
    if not subject:
        return jsonify({'success': False, 'message': '과목을 찾을 수 없습니다.'}), 404
    user = get_current_user()
    if not has_subject_access(subject, user):
        return jsonify({'success': False, 'message': '권한이 없습니다.'}), 403

    email = request.form.get('email')
    role = request.form.get('role', 'student')
    require_approval = request.form.get('require_approval', 'true') == 'true'

    if role not in ['instructor', 'assistant', 'student', 'ta', 'auditor']:
        flash('잘못된 역할입니다.', 'danger')
        return redirect(url_for('subjects.members', subject_id=subject_id))

    target_user = dao.get_user_by_email(email)
    if not target_user:
        flash('해당 이메일의 사용자를 찾을 수 없습니다.', 'danger')
        return redirect(url_for('subjects.members', subject_id=subject_id))

    existing = dao.get_subject_member(subject_id, target_user['uid'])
    if existing:
        dao.update_subject_member(existing['id'], {'role': role})
        flash(f'{target_user.get("display_name", target_user.get("email"))}의 역할이 {get_role_display(role)}(으)로 변경되었습니다.', 'success')
    else:
        subject_enrollment = dao.get_subject_enrollment(subject_id, target_user['uid'])

        enrollment_role = 'student' if role in ['student', 'instructor', 'assistant'] else role

        if subject_enrollment:
            if subject_enrollment.get('status') == 'approved':
                flash('이미 등록된 사용자입니다.', 'info')
                return redirect(url_for('subjects.members', subject_id=subject_id))
            dao.update_subject_enrollment(subject_id, target_user['uid'], {
                'status': 'pending' if require_approval else 'approved',
                'role': enrollment_role
            })
        else:
            dao.create_subject_enrollment({
                'subject_id': subject_id,
                'user_id': target_user['uid'],
                'status': 'pending' if require_approval else 'approved',
                'role': enrollment_role,
                'created_at': datetime.utcnow().isoformat()
            })

        if require_approval:
            dao.create_notification({
                'type': 'enrollment_invite',
                'title': f'{subject["title"]} 과목 등록 초대',
                'message': f'{subject["title"]} 과목에 {get_role_display(enrollment_role)}(으)로 초대되었습니다.',
                'user_id': target_user['uid'],
                'data': {'subject_id': subject_id, 'role': enrollment_role},
                'created_at': datetime.utcnow().isoformat(),
                'is_read': False
            })
            flash(f'{target_user.get("display_name", target_user.get("email"))}에게 등록 초대를 발송했습니다. 승인 대기 중입니다.', 'success')
        else:
            dao.update_subject_enrollment(subject_id, target_user['uid'], {
                'approved_at': datetime.utcnow().isoformat()
            })

            dao.create_subject_member({
                'subject_id': subject_id,
                'user_id': target_user['uid'],
                'role': role
            })

            courses = dao.get_courses_by_subject(subject_id)
            courses = [c for c in courses if not c.get('deleted_at') and c.get('visibility') != 'private']
            for course in courses:
                if not dao.is_enrolled(target_user['uid'], course['id']):
                    dao.create_enrollment({
                        'course_id': course['id'],
                        'user_id': target_user['uid'],
                        'created_at': datetime.utcnow().isoformat()
                    })

            flash(f'{target_user.get("display_name", target_user.get("email"))}이(가) {get_role_display(role)}(으)로 추가되었습니다.', 'success')

    return redirect(url_for('subjects.members', subject_id=subject_id))


@bp.route('/<subject_id>/members/<member_id>/remove', methods=['POST'])
@auth_required
def remove_member(subject_id, member_id):
    subject = dao.get_subject(subject_id)
    if not subject:
        flash('해당 과목을 찾을 수 없습니다.', 'danger')
        return redirect(url_for('subjects.list_subjects'))
    user = get_current_user()
    if not has_subject_access(subject, user):
        flash('권한이 없습니다.', 'danger')
        return redirect(url_for('subjects.members', subject_id=subject_id))

    # Retrieve the member to verify subject_id
    members_list = dao.get_subject_members(subject_id)
    member = None
    for m in members_list:
        if m['id'] == member_id:
            member = m
            break

    if not member:
        flash('잘못된 요청입니다.', 'danger')
        return redirect(url_for('subjects.members', subject_id=subject_id))

    if member['subject_id'] != subject_id:
        flash('잘못된 요청입니다.', 'danger')
        return redirect(url_for('subjects.members', subject_id=subject_id))

    member_user = dao.get_user(member['user_id'])
    user_email = member_user['email'] if member_user else 'unknown'
    dao.delete_subject_member(member_id)
    flash(f'{user_email}이(가) 역할에서 제외되었습니다.', 'success')
    return redirect(url_for('subjects.members', subject_id=subject_id))


@bp.route('/<subject_id>/members/<member_id>/change-role', methods=['POST'])
@auth_required
def change_member_role(subject_id, member_id):
    subject = dao.get_subject(subject_id)
    if not subject:
        return jsonify({'success': False, 'message': '과목을 찾을 수 없습니다.'}), 404
    user = get_current_user()
    if not has_subject_access(subject, user):
        return jsonify({'success': False, 'message': '권한이 없습니다.'}), 403

    # Retrieve the member to verify subject_id
    members_list = dao.get_subject_members(subject_id)
    member = None
    for m in members_list:
        if m['id'] == member_id:
            member = m
            break

    if not member:
        return jsonify({'success': False, 'message': '멤버를 찾을 수 없습니다.'}), 404

    if member['subject_id'] != subject_id:
        return jsonify({'success': False, 'message': '잘못된 요청입니다.'}), 400

    new_role = request.form.get('role')
    if new_role not in ['instructor', 'assistant', 'student']:
        return jsonify({'success': False, 'message': '잘못된 역할입니다.'}), 400

    dao.update_subject_member(member_id, {'role': new_role})
    member_user = dao.get_user(member['user_id'])
    member_email = member_user['email'] if member_user else 'unknown'
    flash(f'{member_email}의 역할이 {get_role_display(new_role)}(으)로 변경되었습니다.', 'success')
    return redirect(url_for('subjects.members', subject_id=subject_id))


@bp.route('/<subject_id>/members/excel-template')
@auth_required
def download_member_template(subject_id):
    from flask import Response
    subject = dao.get_subject(subject_id)
    if not subject:
        flash('해당 과목을 찾을 수 없습니다.', 'danger')
        return redirect(url_for('subjects.list_subjects'))
    user = get_current_user()
    if not has_subject_access(subject, user):
        flash('권한이 없습니다.', 'danger')
        return redirect(url_for('subjects.view', subject_id=subject_id))

    wb = Workbook()
    ws = wb.active
    ws.title = '사용자 등록'

    ws['A1'] = '이메일 (필수)'
    ws['B1'] = '역할 (선택: instructor/assistant/student)'
    ws['C1'] = '이름 (신규 사용자시 필수)'

    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
    for col in ['A', 'B', 'C']:
        ws[f'{col}1'].font = header_font
        ws[f'{col}1'].fill = header_fill
        ws.column_dimensions[col].width = 40

    example_font = Font(italic=True, color='888888')
    ws['A2'] = '(예시) example@email.com'
    ws['B2'] = '(예시) student'
    ws['C2'] = '(예시) 홍길동'
    for col in ['A', 'B', 'C']:
        ws[f'{col}2'].font = example_font

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment;filename=member_template_{subject_id}.xlsx'}
    )


@bp.route('/<subject_id>/members/upload-excel', methods=['POST'])
@auth_required
def upload_members_excel(subject_id):
    subject = dao.get_subject(subject_id)
    if not subject:
        flash('해당 과목을 찾을 수 없습니다.', 'danger')
        return redirect(url_for('subjects.list_subjects'))
    user = get_current_user()
    if not has_subject_access(subject, user):
        flash('권한이 없습니다.', 'danger')
        return redirect(url_for('subjects.members', subject_id=subject_id))

    if 'excel_file' not in request.files:
        flash('파일을 선택해주세요.', 'danger')
        return redirect(url_for('subjects.members', subject_id=subject_id))

    file = request.files['excel_file']
    if file.filename == '':
        flash('파일을 선택해주세요.', 'danger')
        return redirect(url_for('subjects.members', subject_id=subject_id))

    if not file.filename.endswith(('.xlsx', '.xls')):
        flash('엑셀 파일(.xlsx, .xls)만 업로드 가능합니다.', 'danger')
        return redirect(url_for('subjects.members', subject_id=subject_id))

    try:
        wb = load_workbook(file)
        ws = wb.active

        added_count = 0
        updated_count = 0
        skipped_count = 0
        error_messages = []

        for row_num, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            if not row or not row[0]:
                continue

            email = str(row[0]).strip()
            role = str(row[1]).strip().lower() if row[1] else 'student'
            name = str(row[2]).strip() if len(row) > 2 and row[2] else None

            import re
            email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
            if not re.match(email_pattern, email):
                skipped_count += 1
                error_messages.append(f'행 {row_num}: {email} - 잘못된 이메일 형식')
                continue

            if role not in ['instructor', 'assistant', 'student']:
                role = 'student'

            target_user = dao.get_user_by_email(email)
            if not target_user:
                if not name:
                    skipped_count += 1
                    error_messages.append(f'행 {row_num}: {email} - 등록되지 않은 사용자 (신규 등록시 이름 필수)')
                    continue

                import secrets
                from app.firebase_init import get_auth
                auth = get_auth()

                temp_password = secrets.token_urlsafe(8)
                try:
                    user_record = auth.create_user(email=email, password=temp_password)
                    dao.create_user(user_record.uid, {
                        'email': email,
                        'display_name': name,
                        'full_name': name,
                        'role': 'student' if role == 'student' else 'instructor',
                        'created_at': datetime.utcnow().isoformat()
                    })
                    target_user = dao.get_user(user_record.uid)
                    error_messages.append(f'행 {row_num}: {email} - 신규 사용자 생성 (임시 비밀번호: {temp_password})')
                except Exception as e:
                    skipped_count += 1
                    error_messages.append(f'행 {row_num}: {email} - 사용자 생성 실패: {str(e)}')
                    continue

            existing = dao.get_subject_member(subject_id, target_user['uid'])
            if existing:
                if existing['role'] != role:
                    dao.update_subject_member(existing['id'], {'role': role})
                    updated_count += 1
                else:
                    skipped_count += 1
            else:
                dao.create_subject_member({
                    'subject_id': subject_id,
                    'user_id': target_user['uid'],
                    'role': role
                })

                subject_enrollment = dao.get_subject_enrollment(subject_id, target_user['uid'])
                if not subject_enrollment:
                    dao.create_subject_enrollment({
                        'subject_id': subject_id,
                        'user_id': target_user['uid'],
                        'status': 'approved',
                        'created_at': datetime.utcnow().isoformat()
                    })

                if role == 'student':
                    courses = dao.get_courses_by_subject(subject_id)
                    courses = [c for c in courses if not c.get('deleted_at') and c.get('visibility') != 'private']
                    for course in courses:
                        if not dao.is_enrolled(target_user['uid'], course['id']):
                            dao.create_enrollment({
                                'course_id': course['id'],
                                'user_id': target_user['uid'],
                                'created_at': datetime.utcnow().isoformat()
                            })

                added_count += 1

        message_parts = []
        if added_count > 0:
            message_parts.append(f'{added_count}명 추가')
        if updated_count > 0:
            message_parts.append(f'{updated_count}명 역할 변경')
        if skipped_count > 0:
            message_parts.append(f'{skipped_count}명 건너뜀')

        if message_parts:
            flash(f'엑셀 업로드 완료: {", ".join(message_parts)}', 'success')
        else:
            flash('처리할 데이터가 없습니다.', 'info')

        if error_messages and len(error_messages) <= 5:
            for msg in error_messages:
                flash(msg, 'warning')
        elif error_messages:
            flash(f'{len(error_messages)}개의 행에서 오류가 발생했습니다.', 'warning')

    except Exception as e:
        flash(f'파일 처리 중 오류가 발생했습니다: {str(e)}', 'danger')

    return redirect(url_for('subjects.members', subject_id=subject_id))


@bp.route('/<subject_id>/enrollment/approve', methods=['POST'])
@auth_required
def approve_enrollment(subject_id):
    subject = dao.get_subject(subject_id)
    if not subject:
        flash('해당 과목을 찾을 수 없습니다.', 'danger')
        return redirect(url_for('main.dashboard'))

    user = get_current_user()

    enrollment = dao.get_subject_enrollment(subject_id, user['uid'])

    if not enrollment or enrollment.get('status') != 'pending':
        flash('대기 중인 등록 요청이 없습니다.', 'warning')
        return redirect(url_for('main.dashboard'))

    dao.update_subject_enrollment(subject_id, user['uid'], {
        'status': 'approved',
        'approved_at': datetime.utcnow().isoformat()
    })

    member = dao.get_subject_member(subject_id, user['uid'])
    if not member:
        member_role = enrollment.get('role', 'student')
        if member_role not in ['ta', 'auditor']:
            member_role = 'student'
        dao.create_subject_member({
            'subject_id': subject_id,
            'user_id': user['uid'],
            'role': member_role
        })

    courses = dao.get_courses_by_subject(subject_id)
    courses = [c for c in courses if not c.get('deleted_at') and c.get('visibility') != 'private']
    for course in courses:
        if not dao.is_enrolled(user['uid'], course['id']):
            dao.create_enrollment({
                'course_id': course['id'],
                'user_id': user['uid'],
                'created_at': datetime.utcnow().isoformat()
            })

    flash(f'{subject["title"]} 과목에 등록되었습니다.', 'success')
    return redirect(url_for('subjects.view', subject_id=subject_id))


@bp.route('/<subject_id>/enrollment/reject', methods=['POST'])
@auth_required
def reject_enrollment(subject_id):
    subject = dao.get_subject(subject_id)
    if not subject:
        flash('해당 과목을 찾을 수 없습니다.', 'danger')
        return redirect(url_for('main.dashboard'))

    user = get_current_user()

    enrollment = dao.get_subject_enrollment(subject_id, user['uid'])

    if not enrollment or enrollment.get('status') != 'pending':
        flash('대기 중인 등록 요청이 없습니다.', 'warning')
        return redirect(url_for('main.dashboard'))

    dao.update_subject_enrollment(subject_id, user['uid'], {
        'status': 'rejected',
        'rejected_at': datetime.utcnow().isoformat()
    })

    flash(f'{subject["title"]} 과목 등록을 거절했습니다.', 'info')
    return redirect(url_for('main.dashboard'))


@bp.route('/<subject_id>/enrollment/<user_id>/admin-approve', methods=['POST'])
@auth_required
def admin_approve_enrollment(subject_id, user_id):
    subject = dao.get_subject(subject_id)
    if not subject:
        flash('해당 과목을 찾을 수 없습니다.', 'danger')
        return redirect(url_for('subjects.list_subjects'))

    user = get_current_user()
    if not has_subject_access(subject, user):
        flash('권한이 없습니다.', 'danger')
        return redirect(url_for('subjects.members', subject_id=subject_id))

    enrollment = dao.get_subject_enrollment(subject_id, user_id)

    if not enrollment or enrollment.get('status') != 'pending':
        flash('대기 중인 등록 요청이 없습니다.', 'warning')
        return redirect(url_for('subjects.members', subject_id=subject_id))

    dao.update_subject_enrollment(subject_id, user_id, {
        'status': 'approved',
        'approved_at': datetime.utcnow().isoformat()
    })

    member = dao.get_subject_member(subject_id, user_id)
    if not member:
        member_role = enrollment.get('role', 'student')
        if member_role not in ['ta', 'auditor']:
            member_role = 'student'
        dao.create_subject_member({
            'subject_id': subject_id,
            'user_id': user_id,
            'role': member_role
        })

    courses = dao.get_courses_by_subject(subject_id)
    courses = [c for c in courses if not c.get('deleted_at') and c.get('visibility') != 'private']
    for course in courses:
        if not dao.is_enrolled(user_id, course['id']):
            dao.create_enrollment({
                'course_id': course['id'],
                'user_id': user_id,
                'created_at': datetime.utcnow().isoformat()
            })

    dao.create_notification({
        'type': 'enrollment_approved',
        'title': f'{subject["title"]} 과목 등록 승인',
        'message': f'{subject["title"]} 과목에 {get_role_display(enrollment.get("role", "student"))}(으)로 등록이 승인되었습니다.',
        'user_id': user_id,
        'data': {'subject_id': subject_id, 'role': enrollment.get('role', 'student')},
        'created_at': datetime.utcnow().isoformat(),
        'is_read': False
    })

    target_user = dao.get_user(user_id)
    display_name = target_user.get('display_name', target_user.get('email', 'unknown')) if target_user else 'unknown'
    flash(f'{display_name}의 등록 신청을 승인했습니다.', 'success')
    return redirect(url_for('subjects.members', subject_id=subject_id))


@bp.route('/<subject_id>/enrollment/<user_id>/admin-reject', methods=['POST'])
@auth_required
def admin_reject_enrollment(subject_id, user_id):
    subject = dao.get_subject(subject_id)
    if not subject:
        flash('해당 과목을 찾을 수 없습니다.', 'danger')
        return redirect(url_for('subjects.list_subjects'))

    user = get_current_user()
    if not has_subject_access(subject, user):
        flash('권한이 없습니다.', 'danger')
        return redirect(url_for('subjects.members', subject_id=subject_id))

    enrollment = dao.get_subject_enrollment(subject_id, user_id)

    if not enrollment or enrollment.get('status') != 'pending':
        flash('대기 중인 등록 요청이 없습니다.', 'warning')
        return redirect(url_for('subjects.members', subject_id=subject_id))

    dao.update_subject_enrollment(subject_id, user_id, {
        'status': 'rejected',
        'rejected_at': datetime.utcnow().isoformat()
    })

    dao.create_notification({
        'type': 'enrollment_rejected',
        'title': f'{subject["title"]} 과목 등록 거절',
        'message': f'{subject["title"]} 과목에 {get_role_display(enrollment.get("role", "student"))}(으)로 등록이 거절되었습니다.',
        'user_id': user_id,
        'data': {'subject_id': subject_id, 'role': enrollment.get('role', 'student')},
        'created_at': datetime.utcnow().isoformat(),
        'is_read': False
    })

    target_user = dao.get_user(user_id)
    display_name = target_user.get('display_name', target_user.get('email', 'unknown')) if target_user else 'unknown'
    flash(f'{display_name}의 등록 신청을 거절했습니다.', 'info')
    return redirect(url_for('subjects.members', subject_id=subject_id))


@bp.route('/<subject_id>/change-enrollment-status/<user_id>', methods=['POST'])
@auth_required
def change_enrollment_status(subject_id, user_id):
    subject = dao.get_subject(subject_id)
    if not subject:
        flash('해당 과목을 찾을 수 없습니다.', 'danger')
        return redirect(url_for('subjects.list_subjects'))

    user = get_current_user()
    if not has_subject_access(subject, user):
        flash('권한이 없습니다.', 'danger')
        return redirect(url_for('subjects.members', subject_id=subject_id))

    enrollment = dao.get_subject_enrollment(subject_id, user_id)

    if not enrollment:
        flash('등록 정보를 찾을 수 없습니다.', 'warning')
        return redirect(url_for('subjects.members', subject_id=subject_id))

    new_status = request.form.get('status')
    new_role = request.form.get('role')

    target_user = dao.get_user(user_id)
    display_name = target_user.get('display_name', target_user.get('email', 'unknown')) if target_user else 'unknown'

    update_data = {}

    if new_status and new_status in ['approved', 'pending', 'rejected']:
        update_data['status'] = new_status
        if new_status == 'rejected':
            update_data['rejected_at'] = datetime.utcnow().isoformat()
        elif new_status == 'approved' and enrollment.get('status') != 'approved':
            update_data['rejected_at'] = None
        flash(f'{display_name}의 상태가 변경되었습니다.', 'success')

    if new_role and new_role in ['student', 'ta']:
        update_data['role'] = new_role
        flash(f'{display_name}의 역할이 변경되었습니다.', 'success')

    if update_data:
        dao.update_subject_enrollment(subject_id, user_id, update_data)

    return redirect(url_for('subjects.members', subject_id=subject_id))


@bp.route('/my-pending-enrollments')
@auth_required
def my_pending_enrollments():
    user = get_current_user()
    pending_enrollments = dao.get_subject_enrollments_by_user(user['uid'], status='pending')

    # Enrich with subject data
    for enrollment in pending_enrollments:
        enrollment['subject'] = dao.get_subject(enrollment['subject_id'])

    return render_template('subjects/pending_enrollments.html',
                          pending_enrollments=pending_enrollments)


@bp.route('/apply-role', methods=['POST'])
@auth_required
def apply_role():
    user = get_current_user()
    subject_id = request.form.get('subject_id')
    role = request.form.get('role')

    if role not in ['ta', 'auditor']:
        flash('잘못된 역할입니다.', 'danger')
        return redirect(url_for('main.dashboard'))

    subject = dao.get_subject(subject_id)
    if not subject:
        flash('해당 과목을 찾을 수 없습니다.', 'danger')
        return redirect(url_for('main.dashboard'))

    existing = dao.get_subject_enrollment(subject_id, user['uid'])

    if existing and existing.get('status') == 'approved':
        flash('이미 해당 과목에 등록되어 있습니다.', 'info')
        return redirect(url_for('subjects.view', subject_id=subject_id))

    if existing and existing.get('status') == 'pending':
        flash('이미 등록 요청이 대기 중입니다.', 'info')
        return redirect(url_for('main.dashboard'))

    if existing:
        dao.update_subject_enrollment(subject_id, user['uid'], {
            'status': 'pending',
            'role': role
        })
    else:
        dao.create_subject_enrollment({
            'subject_id': subject_id,
            'user_id': user['uid'],
            'status': 'pending',
            'role': role,
            'created_at': datetime.utcnow().isoformat()
        })

    dao.create_notification({
        'type': 'role_application',
        'title': f'{get_role_display(role)} 신청',
        'message': f'{user.get("display_name", user.get("email", ""))}님이 {subject["title"]} 과목에 {get_role_display(role)}(으)로 신청했습니다.',
        'user_id': subject['instructor_id'],
        'data': {'subject_id': subject_id, 'user_id': user['uid'], 'role': role},
        'created_at': datetime.utcnow().isoformat(),
        'is_read': False
    })

    flash(f'{subject["title"]} 과목에 {get_role_display(role)}(으)로 신청했습니다. 승인을 기다려주세요.', 'success')
    return redirect(url_for('main.dashboard'))
